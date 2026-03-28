"""
File content extractors for PDF and Excel formats.
Each extractor takes raw file bytes and returns plain text
suitable for the ingestion pipeline.
"""

from __future__ import annotations

import io
import logging
from enum import Enum

import openpyxl
import pdfplumber
import xlrd

logger = logging.getLogger(__name__)


class FileType(str, Enum):
    PDF = "pdf"
    XLSX = "xlsx"
    XLS = "xls"


MIME_TO_FILETYPE: dict[str, FileType] = {
    "application/pdf": FileType.PDF,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": FileType.XLSX,
    "application/vnd.ms-excel": FileType.XLS,
}

EXT_TO_FILETYPE: dict[str, FileType] = {
    ".pdf": FileType.PDF,
    ".xlsx": FileType.XLSX,
    ".xls": FileType.XLS,
}


def detect_file_type(filename: str, content_type: str | None) -> FileType:
    """
    Determine file type from MIME type first, then fall back to extension.
    Raises ValueError if the file type is not supported.
    """
    if content_type and content_type in MIME_TO_FILETYPE:
        return MIME_TO_FILETYPE[content_type]

    suffix = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix in EXT_TO_FILETYPE:
        return EXT_TO_FILETYPE[suffix]

    raise ValueError(
        f"Unsupported file type: '{filename}' (content_type={content_type}). "
        f"Supported formats: PDF, XLSX, XLS."
    )


def extract_pdf_text(file_bytes: bytes) -> str:
    """
    Extract text from a PDF file, preserving table structure where possible.
    """
    text_parts: list[str] = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        cleaned = [str(cell).strip() if cell else "" for cell in row]
                        text_parts.append(" | ".join(cleaned))
                    text_parts.append("")

            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)

            text_parts.append(f"\n--- Page {page_num} ---\n")

    result = "\n".join(text_parts).strip()
    if not result or len(result) < 10:
        raise ValueError(
            "Could not extract meaningful text from PDF. "
            "The file may be image-based (scanned). OCR is not supported."
        )
    return result


def extract_xlsx_text(file_bytes: bytes) -> str:
    """
    Extract tabular data from an .xlsx file.
    Each sheet becomes a section with rows rendered as pipe-delimited text.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    text_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"=== Sheet: {sheet_name} ===\n")

        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                text_parts.append(" | ".join(cells))

        text_parts.append("")

    wb.close()
    result = "\n".join(text_parts).strip()
    if not result or result.replace("=", "").replace("Sheet:", "").strip() == "":
        raise ValueError("Excel file contains no data.")
    return result


def extract_xls_text(file_bytes: bytes) -> str:
    """
    Extract tabular data from a legacy .xls file using xlrd.
    """
    wb = xlrd.open_workbook(file_contents=file_bytes)
    text_parts: list[str] = []

    for sheet in wb.sheets():
        text_parts.append(f"=== Sheet: {sheet.name} ===\n")

        for row_idx in range(sheet.nrows):
            cells = [str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
            if any(c.strip() for c in cells):
                text_parts.append(" | ".join(cells))

        text_parts.append("")

    result = "\n".join(text_parts).strip()
    if not result or result.replace("=", "").replace("Sheet:", "").strip() == "":
        raise ValueError("Excel file contains no data.")
    return result


EXTRACTORS = {
    FileType.PDF: extract_pdf_text,
    FileType.XLSX: extract_xlsx_text,
    FileType.XLS: extract_xls_text,
}


def extract_text(file_bytes: bytes, file_type: FileType) -> str:
    """Route to the correct extractor based on file type."""
    return EXTRACTORS[file_type](file_bytes)
